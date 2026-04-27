"""
Build COUNTRY_CONTEXT — one row per origin country with the four
external context fields used by the Atlas country panel:

  Freedom House status   (Free / Partly Free / Not Free + 0–100 score)
  UCDP conflict events   (event count + battle-related deaths, latest year)
  UNHCR people of concern (refugees + asylum seekers + IDPs from origin)
  GDP per capita PPP     (latest non-null year, USD)

Inputs (caches, populated by the matching fetch_*.py scripts):
  cache/worldbank/ny_gdp_pcap_pp_cd.json
  cache/freedomhouse/All_data_FIW.xlsx
  cache/unhcr/population/<year>.json
  cache/ucdp/ged_<version>_<year>.json   (optional; needs UCDP_API_TOKEN)

Output:
  data/country-context-data.js   — sets window.COUNTRY_CONTEXT and
                                    window.COUNTRY_CONTEXT_META.

Each entry is keyed by ISO 3166 alpha-3. Sources that are missing or
empty are simply absent from the entry; the consumer renders only the
fields that are present.

Usage:
    python scripts/build_country_context.py
"""
from __future__ import annotations

import argparse
import datetime as dt
import json
import re
import sys
from collections import defaultdict
from pathlib import Path

import openpyxl
import pycountry


def _load_worldbank(cache_dir: Path) -> tuple[dict[str, dict], str | None]:
    f = cache_dir / "worldbank" / "ny_gdp_pcap_pp_cd.json"
    if not f.exists():
        return {}, None
    payload = json.loads(f.read_text(encoding="utf-8"))
    rows = payload.get("rows", [])
    by_iso: dict[str, dict] = {}
    for r in rows:
        iso = (r.get("countryiso3code") or "").strip()
        v = r.get("value")
        date = r.get("date")
        if not iso or len(iso) != 3 or v is None or date is None:
            continue  # skip aggregate-region rows ("WLD" length 3 but actually used) — we'll guard with iso list below
        # Take the latest non-null per country.
        cur = by_iso.get(iso)
        if cur is None or int(date) > cur["year"]:
            by_iso[iso] = {"value": float(v), "year": int(date)}
    # Drop World Bank aggregate ISO3-like codes that aren't real countries.
    AGG = {"WLD", "EUU", "EMU", "OED", "IBT", "LDC", "LIC", "LMY", "LMC", "UMC",
           "HIC", "MIC", "INX", "MEA", "PRE", "SST", "EAS", "EAP", "ECS", "ECA",
           "FCS", "HPC", "IDA", "IDB", "IDX", "LAC", "LCN", "LTE", "MNA", "NAC",
           "OSS", "PSS", "PST", "SAS", "SSA", "SSF", "TEA", "TEC", "TLA", "TMN",
           "TSA", "TSS", "ARB", "CEB", "CSS", "AFE", "AFW", "EAR"}
    return ({iso: v for iso, v in by_iso.items() if iso not in AGG}, payload.get("fetched_at"))


_FH_NAME_OVERRIDES = {
    "Bahamas": "BHS", "Bolivia": "BOL", "Brunei": "BRN", "Cape Verde": "CPV",
    "Czech Republic": "CZE", "Czechia": "CZE", "Congo (Brazzaville)": "COG",
    "Congo (Kinshasa)": "COD", "Democratic Republic of Congo": "COD",
    "East Timor": "TLS", "Timor-Leste": "TLS",
    "Iran": "IRN", "Ivory Coast": "CIV", "Côte d'Ivoire": "CIV",
    "Laos": "LAO", "Macedonia": "MKD", "North Macedonia": "MKD",
    "Micronesia": "FSM", "Moldova": "MDA", "Russia": "RUS",
    "South Korea": "KOR", "North Korea": "PRK", "Korea, North": "PRK",
    "Korea, South": "KOR", "Republic of Korea": "KOR",
    "Burma": "MMR", "Myanmar": "MMR",
    "Saint Kitts and Nevis": "KNA", "Saint Lucia": "LCA",
    "Saint Vincent and the Grenadines": "VCT", "Sao Tome and Principe": "STP",
    "São Tomé and Príncipe": "STP",
    "Syria": "SYR", "Taiwan": "TWN", "Tanzania": "TZA",
    "United States": "USA", "United Kingdom": "GBR",
    "Vatican City": "VAT", "Holy See": "VAT",
    "Venezuela": "VEN", "Vietnam": "VNM",
    "Eswatini": "SWZ", "Swaziland": "SWZ",
    "Cabo Verde": "CPV", "Turkey": "TUR", "Türkiye": "TUR",
    # Common Freedom House territory rows we deliberately skip:
}
# Territory rows we don't try to map to an ISO3 (FIW lists them as 't' anyway).
_FH_SKIP = {
    "Abkhazia", "Crimea", "Eastern Donbas", "Gaza Strip", "Hong Kong", "Indian Kashmir",
    "Kosovo", "Nagorno-Karabakh", "Northern Cyprus", "Pakistani Kashmir",
    "Puerto Rico", "Somaliland", "South Ossetia", "Tibet", "Transnistria",
    "Western Sahara", "West Bank", "West Papua", "Northern Ireland",
}


def _name_to_iso3(name: str) -> str | None:
    if not name:
        return None
    if name in _FH_SKIP:
        return None
    if name in _FH_NAME_OVERRIDES:
        return _FH_NAME_OVERRIDES[name]
    try:
        c = pycountry.countries.lookup(name)
        return c.alpha_3
    except LookupError:
        # Try common-name match.
        n = name.split(",")[0].strip()
        try:
            return pycountry.countries.lookup(n).alpha_3
        except LookupError:
            return None


def _load_freedomhouse(cache_dir: Path) -> tuple[dict[str, dict], str | None]:
    f = cache_dir / "freedomhouse" / "All_data_FIW.xlsx"
    if not f.exists():
        return {}, None
    wb = openpyxl.load_workbook(f, read_only=True, data_only=True)
    sheet_name = next((s for s in wb.sheetnames if s.lower().startswith("fiw")), wb.sheetnames[0])
    ws = wb[sheet_name]
    # Header on row 2 (1-indexed); data from row 3.
    rows_iter = ws.iter_rows(values_only=True)
    headers: list[str] = []
    for i, row in enumerate(rows_iter):
        if i == 1:
            headers = [str(h or "").strip() for h in row]
            break
    col = {h: idx for idx, h in enumerate(headers)}
    needed = {"Country/Territory", "C/T", "Edition", "Status", "PR rating", "CL rating"}
    if not needed.issubset(col.keys()):
        raise RuntimeError(f"Freedom House sheet missing expected columns; got {headers[:12]}")

    by_iso: dict[str, dict] = {}
    unmapped = 0
    for row in rows_iter:
        name = row[col["Country/Territory"]]
        ct = row[col["C/T"]]
        if not name or ct != "c":
            continue
        edition = row[col["Edition"]]
        status = row[col["Status"]]
        pr = row[col["PR rating"]]
        cl = row[col["CL rating"]]
        if edition is None or status is None:
            continue
        iso = _name_to_iso3(str(name))
        if not iso:
            unmapped += 1
            continue
        cur = by_iso.get(iso)
        if cur is None or int(edition) > cur["year"]:
            # Combine PR (1-7) and CL (1-7) into a 0-100 score, lower-is-better
            # → flip and rescale so 100 = most free, 0 = least free.
            score = None
            if pr is not None and cl is not None:
                score = round((1 - ((float(pr) + float(cl)) / 2 - 1) / 6) * 100)
            by_iso[iso] = {
                "status": _expand_status(str(status)),
                "score": score,
                "year": int(edition),
            }
    if unmapped:
        print(f"  Freedom House: {unmapped} rows had no ISO3 match (mostly territories) — skipped", file=sys.stderr)
    # Source URL (if recorded by the fetcher).
    src_url_p = cache_dir / "freedomhouse" / "source_url.txt"
    src = src_url_p.read_text(encoding="utf-8").strip() if src_url_p.exists() else None
    return by_iso, src


def _expand_status(short: str) -> str:
    return {"F": "Free", "PF": "Partly Free", "NF": "Not Free"}.get(short.strip().upper(), short)


def _load_unhcr(cache_dir: Path) -> tuple[dict[str, dict], int | None]:
    pop_dir = cache_dir / "unhcr" / "population"
    if not pop_dir.exists():
        return {}, None
    # Find newest year file.
    files = sorted(pop_dir.glob("*.json"))
    if not files:
        return {}, None
    by_iso: dict[str, dict] = {}
    latest_year = None
    for f in files:
        payload = json.loads(f.read_text(encoding="utf-8"))
        items = payload.get("items", [])
        year = payload.get("year")
        if not items:
            continue
        for it in items:
            iso = (it.get("coo") or it.get("coo_iso") or "").strip()
            if not iso or len(iso) != 3:
                continue
            # Sum the persons-of-concern dimensions; treat '-' as 0.
            def n(v):
                if v in (None, "-", ""):
                    return 0
                try:
                    return int(v)
                except Exception:
                    return 0
            total = (n(it.get("refugees")) + n(it.get("asylum_seekers"))
                     + n(it.get("idps")) + n(it.get("ooc")) + n(it.get("stateless")))
            cur = by_iso.get(iso)
            if cur is None or year > cur["year"]:
                by_iso[iso] = {
                    "year": year,
                    "total": total,
                    "refugees": n(it.get("refugees")),
                    "asylum_seekers": n(it.get("asylum_seekers")),
                    "idps": n(it.get("idps")),
                }
            if latest_year is None or (year is not None and year > latest_year):
                latest_year = year
    return by_iso, latest_year


def _load_ucdp(cache_dir: Path) -> dict[str, dict]:
    ucdp_dir = cache_dir / "ucdp"
    if not ucdp_dir.exists():
        return {}
    files = sorted(ucdp_dir.glob("ged_*.json"))
    if not files:
        return {}
    by_iso: dict[str, dict] = defaultdict(lambda: {"events": 0, "deaths": 0, "year": None})
    for f in files:
        payload = json.loads(f.read_text(encoding="utf-8"))
        year = payload.get("year")
        for ev in payload.get("events", []):
            iso = (ev.get("country_id") or "").strip()  # numeric? UCDP uses numeric Gleditsch-Ward
            # UCDP events also carry 'country' as ISO3-ish. Try several keys.
            iso3 = (ev.get("country") or ev.get("iso3") or "").strip().upper()
            if not iso3 or len(iso3) != 3:
                continue
            ent = by_iso[iso3]
            ent["events"] += 1
            ent["deaths"] += int(ev.get("best") or 0)
            if ent["year"] is None or (year and year > ent["year"]):
                ent["year"] = year
    return dict(by_iso)


def main() -> int:
    p = argparse.ArgumentParser(description=__doc__.split("\n\n")[0])
    p.add_argument("--cache-dir", default="cache")
    p.add_argument("--out", default="data/country-context-data.js")
    args = p.parse_args()

    cache_root = Path(args.cache_dir)
    out_p = Path(args.out)
    out_p.parent.mkdir(parents=True, exist_ok=True)

    wb_data, wb_fetched = _load_worldbank(cache_root)
    fh_data, fh_url = _load_freedomhouse(cache_root)
    unhcr_data, unhcr_latest_year = _load_unhcr(cache_root)
    ucdp_data = _load_ucdp(cache_root)

    # Union of all ISO3 keys — anything that has at least one source contributes.
    isos = set(wb_data) | set(fh_data) | set(unhcr_data) | set(ucdp_data)
    out: dict[str, dict] = {}
    for iso in sorted(isos):
        entry: dict[str, object] = {}
        if iso in fh_data:
            fh = fh_data[iso]
            entry["freedomHouse"] = {
                "status": fh["status"],
                "score": fh["score"],
                "year": fh["year"],
            }
        if iso in ucdp_data:
            u = ucdp_data[iso]
            entry["ucdp"] = {
                "events": u["events"],
                "deaths": u["deaths"],
                "year": u["year"],
            }
        if iso in unhcr_data:
            uh = unhcr_data[iso]
            entry["unhcr"] = {
                "total": uh["total"],
                "refugees": uh["refugees"],
                "asylum_seekers": uh["asylum_seekers"],
                "idps": uh["idps"],
                "year": uh["year"],
            }
        if iso in wb_data:
            w = wb_data[iso]
            entry["gdpPerCapitaPPP"] = {
                "value": round(w["value"], 1),
                "year": w["year"],
            }
        if entry:
            out[iso] = entry

    sources_used = []
    if fh_data: sources_used.append("Freedom House FIW")
    if wb_data: sources_used.append("World Bank — GDP per capita PPP")
    if unhcr_data: sources_used.append("UNHCR — persons of concern")
    if ucdp_data: sources_used.append("UCDP — GED events")

    meta = {
        "generatedAt": dt.datetime.now(dt.timezone.utc).isoformat(),
        "sources": sources_used,
        "freedomHouseUrl": fh_url,
        "unhcrLatestYear": unhcr_latest_year,
        "ucdpAvailable": bool(ucdp_data),
        "countryCount": len(out),
    }

    js = (
        "/* AUTO-GENERATED by scripts/build_country_context.py. Do not edit. */\n"
        f"window.COUNTRY_CONTEXT = {json.dumps(out, ensure_ascii=False)};\n"
        f"window.COUNTRY_CONTEXT_META = {json.dumps(meta, ensure_ascii=False)};\n"
    )
    out_p.write_text(js, encoding="utf-8")
    print(f"wrote {out_p}  ({len(out)} countries; sources: {', '.join(sources_used) or 'none'})")
    return 0


if __name__ == "__main__":
    sys.exit(main())
